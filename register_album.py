import os
import time
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException, NoSuchElementException, UnexpectedAlertPresentException
from openpyxl import load_workbook
import datetime as dt
from selenium.webdriver.common.keys import Keys
import traceback

LOGIN_URL = "https://www.mims.or.kr/login"
ALBUM_REGISTER_URL = "https://www.mims.or.kr/mypage/meta"


def login(driver, mims_id: str, mims_password: str) -> bool:
    driver.get(LOGIN_URL)
    try:
        wait = WebDriverWait(driver, 10)
        id_input = wait.until(EC.presence_of_element_located((By.ID, "inputEmail")))
        pw_input = wait.until(EC.presence_of_element_located((By.ID, "inputPwd")))
        login_button = wait.until(EC.element_to_be_clickable((By.ID, "login-btn")))

        id_input.send_keys(mims_id)
        pw_input.send_keys(mims_password)
        login_button.click()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href="/mypage/album"]'))
        )
        print("로그인 성공!")
        return True
    except Exception as e:
        print(f"로그인 실패: {e}")
        return False


def _get_field_value(driver, field_id: str) -> str:
    try:
        el = driver.find_element(By.ID, field_id)
        tag = el.tag_name.lower()
        if tag in ("input", "textarea"):
            return (el.get_attribute("value") or "").strip()
        if tag == "select":
            try:
                return el.find_element(By.CSS_SELECTOR, "option:checked").text.strip()
            except Exception:
                return ""
        return (el.text or "").strip()
    except Exception:
        return ""


def _check_required_and_go_next(driver) -> None:
    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.ID, "meta-album-form"))
    )
    form = driver.find_element(By.ID, "meta-album-form")
    labels = form.find_elements(By.CSS_SELECTOR, "label.required")

    missing = []
    for lb in labels:
        field_id = lb.get_attribute("for") or ""
        if not field_id:
            continue
        val = _get_field_value(driver, field_id)
        if field_id == "albumTitle" and not val:
            try:
                val = driver.find_element(By.ID, "display_album_title").text.strip()
            except Exception:
                pass
        if not val:
            label_text = (lb.text or field_id).strip()
            missing.append((field_id, label_text))

    if missing:
        print("필수 입력 누락으로 앨범정보에서 이동 중단:")
        for fid, ltxt in missing:
            print(f" - {ltxt} (id={fid}) 비어있음")
        return

    try:
        next_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "meta-next-album-btn"))
        )
        next_btn.click()
        print("곡정보 탭으로 이동 중...")
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "track-list"))
        )
        print("곡정보 탭 로딩 완료")
    except Exception as e:
        print(f"곡정보 이동 실패: {e}")


def _drain_alerts_quick(driver, timeout_sec: float = 0.5) -> None:
    """짧은 총량 제한으로 alert/confirm을 빠르게 소화한다."""
    end = time.time() + timeout_sec
    while time.time() < end:
        try:
            WebDriverWait(driver, 0.2).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            msg = alert.text
            print(f"알림 감지: {msg}")
            alert.accept()
            time.sleep(0.01)
        except TimeoutException:
            # 잠깐 쉬었다가 다시 확인
            time.sleep(0.01)
            continue
        except Exception:
            break

def _check_track_required_and_next(driver) -> None:
    # 곡정보 탭 폼/테이블 대기
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "meta-tracks"))
        )
    except TimeoutException:
        # 이미 다른 탭일 수 있으므로 조용히 리턴
        return

    required_ids = ["diskNo", "trackNo", "trackTitle", "trackArtistName", "trackGenre", "duration"]
    missing = []

    # duration은 hidden #duration 또는 HH/MM/SS 조합 확인
    for fid in required_ids:
        if fid == "duration":
            dur_hidden = _get_field_value(driver, "duration")
            if dur_hidden:
                continue
            hh = _get_field_value(driver, "duration_hh")
            mm = _get_field_value(driver, "duration_mm")
            ss = _get_field_value(driver, "duration_ss")
            if not (hh and mm and ss):
                missing.append((fid, "재생시간"))
            continue
        val = _get_field_value(driver, fid)
        if not val:
            label_text = fid
            try:
                label_text = driver.find_element(By.CSS_SELECTOR, f"label[for='{fid}']").text.strip()
            except Exception:
                pass
            missing.append((fid, label_text))

    if missing:
        print("필수 입력 누락으로 곡정보에서 이동 중단:")
        for fid, ltxt in missing:
            print(f" - {ltxt} (id={fid}) 비어있음")
        return

    # 저장 후 다음으로 버튼 클릭
    try:
        save_next_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "update-meta-track-next-track-btn"))
        )
        save_next_btn.click()
        print("곡정보 저장 후 다음으로 이동 중...")
        # 저장/다음 과정에서 alert/confirm이 열릴 수 있으므로 감지 후 모두 확인
        _drain_alerts_quick(driver)

        # 다음 탭(앨범중복확인)에서 버튼/컨텐츠 대기
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "meta-next-right-new-btn"))
            )
            print("앨범중복확인 탭 로딩 완료")
        except TimeoutException:
            try:
                ok_btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='확인' or normalize-space(.)='확 인']"))
                )
                ok_btn.click()
                print("모달 확인 버튼 클릭")
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "meta-next-right-new-btn"))
                )
                print("앨범중복확인 탭 로딩 완료")
            except Exception as e2:
                print(f"앨범중복확인 탭 확인 실패: {e2}")
    except Exception as e:
        print(f"곡정보 저장/이동 실패: {e}")


def _parse_excel_duration_to_hms(value):
    if value is None:
        return None
    try:
        # datetime.time
        if isinstance(value, dt.time):
            return f"{value.hour:02d}", f"{value.minute:02d}", f"{value.second:02d}"
        # datetime.timedelta
        if isinstance(value, dt.timedelta):
            total = int(value.total_seconds())
            hh = total // 3600
            mm = (total % 3600) // 60
            ss = total % 60
            return f"{hh:02d}", f"{mm:02d}", f"{ss:02d}"
        # excel serial number (float or int)
        if isinstance(value, (int, float)):
            # 1 day == 86400 seconds
            total = int(round((float(value) % 1) * 86400))
            hh = total // 3600
            mm = (total % 3600) // 60
            ss = total % 60
            return f"{hh:02d}", f"{mm:02d}", f"{ss:02d}"
        # string like HH:MM:SS or MM:SS
        s = str(value).strip()
        if not s:
            return None
        parts = s.split(":")
        if len(parts) == 3:
            hh, mm, ss = parts
        elif len(parts) == 2:
            hh, mm, ss = "00", parts[0], parts[1]
        else:
            return None
        hh = f"{int(hh):02d}"
        mm = f"{int(mm):02d}"
        ss = f"{int(float(ss)):02d}"
        return hh, mm, ss
    except Exception:
        return None


def _fill_durations_from_excel(driver, excel_path: str) -> None:
    # 트랙 목록 로드 (곡정보 탭 보이게)
    _ensure_tracks_tab(driver)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "track-list")))
    rows = driver.find_elements(By.CSS_SELECTOR, "#track-list tbody tr")
    if not rows:
        print("수록곡 목록을 찾지 못했습니다.")
        return

    # 순서 보존을 위해 import_seq 수집
    row_seqs = [(r.get_attribute("data-import_seq") or "") for r in rows]

    # 엑셀 열기 (첫 시트, N열=14번째, 18행부터)
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 최초 1회: 첫 트랙 활성화 (이후엔 '저장 후 다음으로' 버튼으로만 이동)
    try:
        first_link = WebDriverWait(rows[0], 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "td[for='displayTrackTitle'] a.show-track-btn"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", first_link)
        first_link.click()
        if row_seqs[0]:
            WebDriverWait(driver, 10).until(
                lambda d: d.find_element(By.ID, "importSeq").get_attribute("value") == row_seqs[0]
            )
    except Exception:
        pass

    updated = 0
    total = len(rows)
    for idx in range(total):
        # 현재 트랙에 대응하는 엑셀 행에서 HH:MM:SS 읽기
        excel_row = 18 + idx
        cell = ws.cell(row=excel_row, column=14)  # N열
        hms = _parse_excel_duration_to_hms(cell.value)

        try:
            # 입력칸 찾기 (현재 활성 트랙 폼)
            hh_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "duration_hh")))
            mm_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "duration_mm")))
            ss_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "duration_ss")))
            hidden_duration = driver.find_element(By.ID, "duration")

            existing_hh = (hh_input.get_attribute("value") or "").strip()
            existing_mm = (mm_input.get_attribute("value") or "").strip()
            existing_ss = (ss_input.get_attribute("value") or "").strip()
            existing_hidden = (hidden_duration.get_attribute("value") or "").strip()

            already_filled = bool(existing_hidden or (existing_hh and existing_mm and existing_ss))

            if not already_filled and hms:
                hh, mm, ss = hms

                def type_val(el, val):
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    try:
                        el.click()
                    except Exception:
                        pass
                    try:
                        el.clear()
                    except Exception:
                        el.send_keys(Keys.COMMAND, 'a')
                        el.send_keys(Keys.BACK_SPACE)
                    el.send_keys(val)
                    driver.execute_script(
                        "['keyup','change','input'].forEach(e=>arguments[0].dispatchEvent(new Event(e,{bubbles:true})));",
                        el,
                    )

                type_val(hh_input, hh)
                type_val(mm_input, mm)
                type_val(ss_input, ss)

                if not (hidden_duration.get_attribute("value") or "").strip():
                    driver.execute_script(
                        "arguments[0].value = arguments[1]; ['change','input'].forEach(e=>arguments[0].dispatchEvent(new Event(e,{bubbles:true})));",
                        hidden_duration,
                        f"{hh}:{mm}:{ss}",
                    )

                updated += 1
            elif already_filled:
                print("재생시간이 이미 입력되어 있어 건너뜁니다.")
        except Exception as e:
            print(f"{idx+1}번째 트랙 재생시간 처리 중 요소 탐색/입력 실패: {e}")

        # 다음 곡으로 이동: 반드시 '저장 후 다음으로' 버튼 사용
        try:
            save_next_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "update-meta-track-next-track-btn"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", save_next_btn)
            save_next_btn.click()
            print("저장 후 다음으로 클릭")

            # alert/confirm 연속 처리
            _drain_alerts_quick(driver)

            # 마지막이 아니면 다음 트랙으로 이동 완료 대기 (importSeq 변경 확인)
            if idx < total - 1:
                next_seq = row_seqs[idx + 1]
                WebDriverWait(driver, 6).until(
                    lambda d: d.find_element(By.ID, "importSeq").get_attribute("value") == next_seq
                )
            else:
                # 마지막 곡: 모달 확인 버튼이 있다면 처리
                try:
                    ok_btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='확인' or normalize-space(.)='확 인']"))
                    )
                    ok_btn.click()
                    print("모달 '확인' 버튼 클릭")
                except Exception:
                    pass

        except Exception as e:
            print(f"저장 후 다음으로 클릭 실패: {e}")

    print(f"재생시간 입력 완료: {updated}개 트랙")


def _ensure_tracks_tab(driver) -> None:
    try:
        # 이미 트랙 탭이 활성화되어 있으면 통과
        meta_tracks = driver.find_element(By.ID, "meta-tracks")
        if "show" in meta_tracks.get_attribute("class") and "active" in meta_tracks.get_attribute("class"):
            return
    except Exception:
        pass

    # 상단 탭에서 곡정보 탭 클릭 시도
    try:
        tab_btn = driver.find_element(By.CSS_SELECTOR, "#metaTab a[data-target='#meta-tracks']")
        tab_btn.click()
    except Exception:
        # 탭 버튼이 없으면 앨범정보의 '곡정보 이동' 버튼으로 이동
        try:
            next_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.ID, "meta-next-album-btn"))
            )
            next_btn.click()
        except Exception:
            pass

    WebDriverWait(driver, 10).until(
        lambda d: "show" in d.find_element(By.ID, "meta-tracks").get_attribute("class") and "active" in d.find_element(By.ID, "meta-tracks").get_attribute("class")
    )


def _save_next_on_last_track(driver) -> None:
    # 곡정보 탭 보이기 및 트랙 목록 확보
    _ensure_tracks_tab(driver)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "track-list")))
    rows = driver.find_elements(By.CSS_SELECTOR, "#track-list tbody tr")
    if not rows:
        print("수록곡 목록이 비어 있어 저장-다음 수행을 건너뜁니다.")
        return

    last_row = rows[-1]
    try:
        # 마지막 곡 폼 활성화를 위해 제목 링크 클릭
        link = WebDriverWait(last_row, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "td[for='displayTrackTitle'] a.show-track-btn"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
        link.click()

        # 행의 data-import_seq와 폼의 #importSeq 일치 대기
        row_seq = last_row.get_attribute("data-import_seq") or ""
        if row_seq:
            WebDriverWait(driver, 10).until(
                lambda d: d.find_element(By.ID, "importSeq").get_attribute("value") == row_seq
            )
    except Exception as e:
        print(f"마지막 트랙 활성화 실패: {e}")

    # 저장 후 다음으로 버튼 클릭 시도
    try:
        save_next_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "update-meta-track-next-track-btn"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", save_next_btn)
        save_next_btn.click()
        print("마지막 곡에서 '저장 후 다음으로' 클릭")

        # alert/confirm 연속 처리
        try:
            WebDriverWait(driver, 5).until(EC.alert_is_present())
            while True:
                alert = driver.switch_to.alert
                msg = alert.text
                print(f"알림 감지: {msg}")
                alert.accept()
                time.sleep(0.2)
                try:
                    WebDriverWait(driver, 1).until(EC.alert_is_present())
                    continue
                except TimeoutException:
                    break
            print("알림(확인) 모두 처리 완료")
        except TimeoutException:
            pass

        # 모달형 확인 버튼 처리 (있을 경우)
        try:
            ok_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[normalize-space(.)='확인' or normalize-space(.)='확 인']"))
            )
            ok_btn.click()
            print("모달 '확인' 버튼 클릭")
        except Exception:
            pass

    except Exception as e:
        print(f"마지막 곡 저장-다음 클릭 실패: {e}")


def _handle_meta_confirm(driver) -> bool:
    """앨범중복확인 탭에서 검색앨범 수록곡 유무에 따라 진행/중단 결정.

    반환값: True → 다음 단계 진행, False → 중단(수록곡이 1개 이상 발견됨)
    """
    try:
        # 앨범중복확인 탭(메타 확인) 로딩 대기
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "meta-confirm"))
        )
        # 검색앨범 수록곡 테이블 로딩을 잠깐 대기 (비어있을 수 있음)
        time.sleep(0.3)
        rows = driver.find_elements(By.CSS_SELECTOR, "#search-track-data table tbody tr")
        num = len(rows)
        if num <= 0:
            print("검색앨범 수록곡 없음 → 다음 단계로 진행")
            # 다음 버튼(NEW/ DUP 케이스 모두 대응)
            clicked = False
            for btn_id in ("meta-next-right-new-btn", "meta-next-right-dup-btn"):
                try:
                    btn = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.ID, btn_id))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    btn.click()
                    clicked = True
                    break
                except Exception:
                    continue
            if not clicked:
                print("다음 버튼을 찾지 못했습니다. 진행을 중단합니다.")
                return
            _drain_alerts_quick(driver)
            # 권리정보 탭 진입 확인
            try:
                WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.ID, "right-reg-btn"))
                )
                print("권리정보 탭 진입 완료")
            except TimeoutException:
                print("권리정보 탭 진입 확인 실패(타임아웃)")
            return True
        else:
            print(f"검색앨범 수록곡 {num}개 발견 → 화면 유지")
            # 가능하면 곡명 열 출력(3번째 열 또는 .select-title)
            for i, tr in enumerate(rows, start=1):
                title = ""
                try:
                    title = tr.find_element(By.CSS_SELECTOR, ".select-title").text.strip()
                except Exception:
                    try:
                        title = tr.find_elements(By.TAG_NAME, "td")[2].text.strip()
                    except Exception:
                        title = tr.text.strip()
                print(f" - 검색앨범 곡 {i}: {title}")
            # 하나라도 있으면 중단
            return False
    except Exception as e:
        print(f"앨범중복확인 처리 실패: {e}")
        # 실패 시 보수적으로 중단
        return False


def _open_member_search_modals(driver, close_after: bool = True) -> None:
    """권리정보 탭에서 제작회원/유통회원 검색 모달을 순서대로 띄운다.

    close_after=False이면 마지막(유통회원) 모달은 닫지 않고 그대로 둔다.
    첫(제작회원) 모달은 두 번째 클릭을 위해 항상 닫는다.
    """
    try:
        wrapper = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "track-right-apply"))
        )

        def click_btn(selector_css: str, fallback_name=None):
            try:
                btn_el = wrapper.find_element(By.CSS_SELECTOR, selector_css)
            except Exception:
                if fallback_name is None:
                    raise
                btn_el = wrapper.find_element(By.CSS_SELECTOR, f"button.search-group[name='{fallback_name}']")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_el)
            try:
                btn_el.click()
            except ElementClickInterceptedException:
                driver.execute_script("arguments[0].click();", btn_el)
            return btn_el

        def wait_modal_visible():
            return WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.ID, "rightModal"))
            )

        def close_modal_if_open():
            try:
                modal = driver.find_element(By.ID, "rightModal")
                if modal.is_displayed():
                    # footer Close 우선, 없으면 X
                    try:
                        close_btn = modal.find_element(By.CSS_SELECTOR, ".modal-footer button")
                    except Exception:
                        try:
                            close_btn = modal.find_element(By.CSS_SELECTOR, ".close")
                        except Exception:
                            close_btn = None
                    if close_btn:
                        close_btn.click()
                        WebDriverWait(driver, 3).until(
                            EC.invisibility_of_element_located((By.ID, "rightModal"))
                        )
                        # backdrop 사라짐도 잠깐 확인
                        try:
                            WebDriverWait(driver, 2).until(
                                EC.invisibility_of_element_located((By.CSS_SELECTOR, ".modal-backdrop"))
                            )
                        except TimeoutException:
                            pass
                        print("모달 닫기 완료")
            except Exception:
                pass

        # 1) 제작회원(P) 열기 → 로그 → 반드시 닫기
        click_btn("button.search-group[data-group-category='P']", "search-grup1")
        wait_modal_visible()
        print("제작회원 검색 모달 열림")
        close_modal_if_open()

        # 2) 유통회원(S) 열기 → 로그 → close_after에 따라 닫기/유지
        click_btn("button.search-group[data-group-category='S']", "search-grup2")
        wait_modal_visible()
        print("유통회원 검색 모달 열림")
        if close_after:
            close_modal_if_open()
    except Exception as e:
        print(f"회원 검색 모달 열기 실패: {e}")


# ===== main.py에서 통합: 앨범 찾기 및 ISRC/UCI 발급/추출 유틸 =====
def find_approved_albums(driver):
    try:
        my_album_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[href="/mypage/album"]'))
        )
        my_album_link.click()

        album_container_selector = "div.mims-pmb"
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, album_container_selector))
        )

        approved_albums = []
        album_cards = driver.find_elements(By.CSS_SELECTOR, f"{album_container_selector} .thumbnail-style")
        print(f"현재 페이지에서 {len(album_cards)}개의 앨범을 찾았습니다.")

        for card in album_cards:
            try:
                title_element = card.find_element(By.CSS_SELECTOR, "h3 > a")
                album_title = title_element.text
                album_link_element = card.find_element(By.CSS_SELECTOR, "a.go-view")
                album_code = album_link_element.get_attribute('data-album-code')
                approved_albums.append({"title": album_title, "code": album_code, "element": album_link_element})
                print(f"앨범 찾음: {album_title} (코드: {album_code})")
            except Exception as e:
                print(f"앨범 정보를 가져오는 중 오류 발생: {e}")
        return approved_albums
    except Exception as e:
        print(f"앨범을 찾는 중 오류 발생: {e}")
        return []


def issue_codes(driver):
    print("앨범 상세 페이지로 이동했습니다. ISRC/UCI 코드 확인 및 발급을 시작합니다.")

    def extract_codes(driver):
        codes_list = []
        try:
            track_list_xpath = "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, track_list_xpath)))
            track_rows = driver.find_elements(By.XPATH, track_list_xpath)
            if not track_rows:
                return []
            for row in track_rows:
                isrc_elements = row.find_elements(By.CSS_SELECTOR, "span.g-bg-darkred")
                if not isrc_elements:
                    return None
                title = row.find_element(By.CSS_SELECTOR, "td:nth-child(4) a").text
                isrc = isrc_elements[0].get_attribute("data-clipboard-data")
                uci = row.find_element(By.CSS_SELECTOR, "span.g-bg-blue").get_attribute("data-clipboard-data")
                codes_list.append({"title": title, "isrc": isrc, "uci": uci})
            return codes_list
        except Exception:
            return None

    def _count_rows(driver):
        try:
            track_list_xpath = "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, track_list_xpath)))
            return len(driver.find_elements(By.XPATH, track_list_xpath))
        except Exception:
            return 0

    def _wait_for_isrc_applied(driver, expected_rows):
        deadline = time.time() + 30
        while time.time() < deadline:
            try:
                count_now = len(driver.find_elements(By.CSS_SELECTOR, "span.g-bg-darkred[data-clipboard-data]"))
                if count_now >= 1:
                    break
            except UnexpectedAlertPresentException:
                _accept_all_alerts(driver, "ISRC", max_tries=2)
            except Exception:
                pass
            time.sleep(0.2)
        # 전체 행 반영 대기 (소프트)
        deadline2 = time.time() + 20
        while time.time() < deadline2:
            try:
                if len(driver.find_elements(By.CSS_SELECTOR, "span.g-bg-darkred[data-clipboard-data]")) >= expected_rows:
                    return
            except UnexpectedAlertPresentException:
                _accept_all_alerts(driver, "ISRC", max_tries=2)
            except Exception:
                pass
            time.sleep(0.3)

    def _wait_for_uci_applied(driver, expected_rows):
        deadline = time.time() + 30
        while time.time() < deadline:
            try:
                count_now = len(driver.find_elements(By.CSS_SELECTOR, "span.g-bg-blue[data-clipboard-data]"))
                if count_now >= 1:
                    break
            except UnexpectedAlertPresentException:
                print("[UCI WAIT] UnexpectedAlertPresentException during initial wait")
                print(traceback.format_exc())
                _accept_all_alerts(driver, "UCI", max_tries=2)
            except Exception:
                print("[UCI WAIT] Exception during initial wait")
                print(traceback.format_exc())
            time.sleep(0.2)
        deadline2 = time.time() + 20
        while time.time() < deadline2:
            try:
                if len(driver.find_elements(By.CSS_SELECTOR, "span.g-bg-blue[data-clipboard-data]")) >= expected_rows:
                    return
            except UnexpectedAlertPresentException:
                print("[UCI WAIT] UnexpectedAlertPresentException during full-rows wait")
                print(traceback.format_exc())
                _accept_all_alerts(driver, "UCI", max_tries=2)
            except Exception:
                print("[UCI WAIT] Exception during full-rows wait")
                print(traceback.format_exc())
            time.sleep(0.3)

    def _count_isrc_applied(driver):
        return len(driver.find_elements(By.CSS_SELECTOR, "span.g-bg-darkred[data-clipboard-data]"))

    def _count_uci_applied(driver):
        return len(driver.find_elements(By.CSS_SELECTOR, "span.g-bg-blue[data-clipboard-data]"))

    def _accept_all_alerts(driver, label: str, max_tries: int = 5):
        messages = []
        for i in range(1, max_tries + 1):
            try:
                WebDriverWait(driver, 2).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                try:
                    text = alert.text
                    messages.append(text)
                    print(f"[{label} ALERT #{i}] {text}")
                except Exception:
                    pass
                alert.accept()
                time.sleep(0.2)
            except TimeoutException:
                break
            except Exception as e:
                print(f"[{label} ALERT HANDLER ERROR] {e}")
                print(traceback.format_exc())
                break
        return messages

    def _click_and_accept(driver, by, ident, label):
        try:
            btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, ident)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            driver.execute_script("arguments[0].click();", btn)
            alerts = _accept_all_alerts(driver, label, max_tries=5)
            print(f"{label} 발급 확인 완료. 페이지 반영 대기...")
            for _ in range(3):
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]"))
                    )
                    break
                except UnexpectedAlertPresentException:
                    more = _accept_all_alerts(driver, label, max_tries=5)
                    alerts.extend(more)
                except TimeoutException:
                    more = _accept_all_alerts(driver, label, max_tries=1)
                    alerts.extend(more)
            return True, alerts
        except NoSuchElementException as e:
            print(f"{label} 발급 버튼을 찾을 수 없습니다. {e}")
            print(traceback.format_exc())
            return False, []
        except TimeoutException as e:
            print(f"{label} 발급 확인창이 나타나지 않았거나 페이지 로딩에 실패했습니다. {e}")
            print(traceback.format_exc())
            return False, []

    try:
        total_rows = _count_rows(driver)
        if total_rows == 0:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"))
            )
            total_rows = _count_rows(driver)

        isrc_applied = _count_isrc_applied(driver)
        uci_applied = _count_uci_applied(driver)
        need_isrc = isrc_applied < total_rows
        need_uci = uci_applied < total_rows

        if not need_isrc and not need_uci:
            print("ISRC/UCI가 모두 존재합니다. 코드만 추출합니다.")
            return extract_codes(driver)

        did_issue = False
        if need_isrc:
            print("ISRC 발급 버튼 클릭...")
            prev_isrc = _count_isrc_applied(driver)
            ok_isrc, isrc_alerts = _click_and_accept(driver, By.ID, "setTrackIsrc", "ISRC")
            if ok_isrc:
                _wait_for_isrc_applied(driver, total_rows)
                now_isrc = _count_isrc_applied(driver)
                if now_isrc <= prev_isrc and any("오류" in m for m in isrc_alerts):
                    print("ISRC 발급 경고/오류 감지:")
                    for m in isrc_alerts:
                        print(f" - {m}")
                # ISRC 발급 후 1초 대기 + 새로고침 1회 → 알럿 드레인 → 테이블 재등장 대기
                try:
                    time.sleep(1)
                    driver.refresh()
                    _accept_all_alerts(driver, "After ISRC Refresh", max_tries=3)
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"))
                    )
                    print("ISRC 후 새로고침 완료. UCI 발급을 시도합니다.")
                except Exception as _e:
                    print(f"ISRC 후 새로고침 처리 중 경고/오류: {_e}")
            did_issue = True

        if need_uci:
            print("UCI 발급 버튼 클릭...")
            prev_uci = _count_uci_applied(driver)
            try:
                ok_uci, uci_alerts = _click_and_accept(driver, By.ID, "setTrackUCI", "UCI")
            except Exception as e:
                print("UCI 발급 처리 중 예외 발생")
                print(traceback.format_exc())
                ok_uci, uci_alerts = (False, [])
            if ok_uci:
                try:
                    _wait_for_uci_applied(driver, total_rows)
                except Exception:
                    print("UCI 대기 중 예외")
                    print(traceback.format_exc())
                now_uci = _count_uci_applied(driver)
                if now_uci <= prev_uci:
                    print("UCI 발급 결과 반영이 확인되지 않았습니다. 알럿 메시지:")
                    for m in uci_alerts:
                        print(f" - {m}")
                    # 예외가 없어도 진단을 위해 현재 호출 스택 출력
                    try:
                        print("[UCI DIAG] 현재 호출 스택 (format_stack):")
                        print("".join(traceback.format_stack()))
                    except Exception:
                        pass
                # UCI 발급 후 1초 대기 + 새로고침 1회 → 알럿 드레인 → 테이블 재등장 대기 → 즉시 코드 추출 시도
                try:
                    time.sleep(1)
                    driver.refresh()
                    _accept_all_alerts(driver, "After UCI Refresh", max_tries=3)
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"))
                    )
                    quick_codes = extract_codes(driver)
                    if quick_codes:
                        print("UCI 후 새로고침 → 코드 추출 성공.")
                        return quick_codes
                    else:
                        print("UCI 후 새로고침 → 코드 추출 실패, 재시도를 진행합니다.")
                        # 2차 재시도: UCI 버튼 다시 클릭 → 대기 → 1초 후 새로고침 → 즉시 추출 시도
                        try:
                            print("UCI 재시도 진행...")
                            ok_uci2, uci_alerts2 = _click_and_accept(driver, By.ID, "setTrackUCI", "UCI-RETRY")
                            if ok_uci2:
                                _wait_for_uci_applied(driver, total_rows)
                                time.sleep(1)
                                driver.refresh()
                                _accept_all_alerts(driver, "After UCI Retry Refresh", max_tries=3)
                                WebDriverWait(driver, 20).until(
                                    EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"))
                                )
                                retry_codes = extract_codes(driver)
                                if retry_codes:
                                    print("UCI 재시도 후 새로고침 → 코드 추출 성공.")
                                    return retry_codes
                                else:
                                    print("UCI 재시도 후에도 코드 추출 실패")
                                    try:
                                        print("[UCI DIAG] 재시도 후 추출 실패 스택:")
                                        print("".join(traceback.format_stack()))
                                    except Exception:
                                        pass
                        except Exception:
                            print("UCI 재시도 중 예외")
                            print(traceback.format_exc())
                        try:
                            print("[UCI DIAG] 새로고침 후 추출 실패 시 호출 스택:")
                            print("".join(traceback.format_stack()))
                        except Exception:
                            pass
                except Exception as _e:
                    print("UCI 후 새로고침 처리/추출 중 예외")
                    print(traceback.format_exc())
            did_issue = True

        if did_issue:
            time.sleep(1)
            driver.refresh()
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"))
            )
            first_codes = extract_codes(driver)

            driver.refresh()
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"))
            )
            print("코드 발급/이중 새로고침 후 최종 코드 추출을 시도합니다.")
            final_codes = extract_codes(driver)
            return final_codes or first_codes

        return extract_codes(driver)
    except Exception as e:
        print(f"코드 확인/발급 중 예상치 못한 오류 발생: {e}")
        print(traceback.format_exc())
        return None

def _select_producer_member(driver) -> None:
    """권리정보 탭에서 제작회원 검색 모달을 열고 '케이저'로 검색,
    이메일이 metalfocus로 시작하는 결과를 선택한 뒤 모달을 닫는다.
    """
    try:
        # 권리정보 탭을 확실히 활성화
        try:
            rights_tab_btn = driver.find_element(By.CSS_SELECTOR, "#metaTab a[data-target='#meta-right']")
            rights_tab_btn.click()
        except Exception:
            pass
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "track-right-apply"))
        )
        # 1) 제작회원 검색 버튼 클릭 (열기)
        btn = None
        for locator in [
            (By.CSS_SELECTOR, "#track-right-apply button.search-group[data-group-category='P']"),
            (By.CSS_SELECTOR, "#track-right-apply button.search-group[name='search-grup1']"),
            (By.XPATH, "//div[@id='track-right-apply']//label[contains(.,'제작회원')]/following::button[contains(@class,'search-group')][1]")
        ]:
            try:
                btn = driver.find_element(*locator)
                if btn:
                    break
            except Exception:
                continue
        if btn is None:
            raise Exception("제작회원 검색 버튼을 찾을 수 없습니다.")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        try:
            btn.click()
        except ElementClickInterceptedException:
            driver.execute_script("arguments[0].click();", btn)

        # 2) 모달 표시 대기
        modal = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.ID, "rightModal"))
        )
        print("제작회원 검색 모달 열림")

        # 3) 검색어 입력 ('케이저') - 가능한 경우 #searchValue 사용
        try:
            keyword_input = modal.find_element(By.CSS_SELECTOR, "#searchValue")
        except Exception:
            try:
                keyword_input = modal.find_element(By.CSS_SELECTOR, "input[type='text']")
            except Exception:
                keyword_input = modal.find_element(By.XPATH, ".//input[contains(translate(@name,'KEYWORD','keyword'),'keyword')]")
        driver.execute_script("arguments[0].focus();", keyword_input)
        try:
            keyword_input.clear()
        except Exception:
            pass
        keyword_input.send_keys("케이저")

        # 4) '검색' 버튼 클릭
        # 검색 버튼: #modal-right-search 우선 사용
        search_btn = None
        try:
            search_btn = modal.find_element(By.CSS_SELECTOR, "#modal-right-search")
        except Exception:
            try:
                search_btn = modal.find_element(By.XPATH, ".//button[normalize-space(text())='검색']")
            except Exception:
                try:
                    search_btn = modal.find_element(By.CSS_SELECTOR, "button.btn.btn-primary")
                except Exception:
                    btns = modal.find_elements(By.TAG_NAME, "button")
                    search_btn = btns[0] if btns else None
        if search_btn is None:
            raise Exception("검색 버튼을 찾을 수 없습니다.")
        try:
            search_btn.click()
        except ElementClickInterceptedException:
            driver.execute_script("arguments[0].click();", search_btn)

        # 5) 결과 테이블 로딩 대기 후 대상 선택
        # 일반적으로 모달 본문 내 table tbody tr 구조 가정
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#rightModal #search-right-list tbody tr, #rightModal .modal-body table tbody tr"))
        )
        rows = modal.find_elements(By.CSS_SELECTOR, "#search-right-list tbody tr, .modal-body table tbody tr")
        if not rows:
            print("검색 결과가 없습니다.")
        target_row = None
        for tr in rows:
            tds = tr.find_elements(By.TAG_NAME, "td")
            joined_text = " ".join([td.text.strip() for td in tds]).strip()
            if joined_text:
                # 이메일이 'metalfocus'로 시작하는지 검사 (대소문자 무시)
                if any((cell.text.strip().lower().startswith("metalfocus")) for cell in tds):
                    target_row = tr
                    break
                # 혹시 전체 합친 문자열에도 포함될 수 있으니 보조 체크
                if "metalfocus" in joined_text.lower():
                    target_row = tr
                    break

        if target_row is None:
            raise Exception("'metalfocus'로 시작하는 이메일 결과를 찾지 못했습니다.")

        # 행 안의 '선택' 유사 버튼 우선 클릭, 없으면 행 자체 클릭
        # .select-right 링크 우선 클릭, 없으면 대안 요소 클릭
        clickable = None
        try:
            clickable = target_row.find_element(By.CSS_SELECTOR, ".select-right")
        except Exception:
            try:
                clickable = target_row.find_element(By.XPATH, ".//button[contains(.,'선택') or contains(.,'선 정') or contains(.,'선 택')]")
            except Exception:
                for sel in ["input[type='radio']", "input[type='checkbox']", "a", "button", "td"]:
                    try:
                        clickable = target_row.find_element(By.CSS_SELECTOR, sel)
                        break
                    except Exception:
                        continue
        if clickable is None:
            clickable = target_row
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", clickable)
        try:
            clickable.click()
        except ElementClickInterceptedException:
            driver.execute_script("arguments[0].click();", clickable)
        print("제작회원: metalfocus* 항목 선택 완료")

        # 6) 선택 시 자동 닫힘을 기다림 (모달/백드롭/바디 클래스)
        try:
            WebDriverWait(driver, 6).until(
                EC.invisibility_of_element_located((By.ID, "rightModal"))
            )
            WebDriverWait(driver, 3).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, ".modal-backdrop"))
            )
            WebDriverWait(driver, 3).until(
                lambda d: "modal-open" not in d.find_element(By.TAG_NAME, "body").get_attribute("class")
            )
        except Exception:
            pass
    except Exception as e:
        print(f"제작회원 자동 선택 실패: {e}")


def _select_distributor_member(driver) -> None:
    """권리정보 탭에서 유통회원 검색 모달을 열고 '케이저'로 검색,
    이메일이 metalfocus로 시작하는 결과를 선택한 뒤 모달이 자동으로 닫히는 것을 대기한다.
    """
    try:
        # 권리정보 탭을 확실히 활성화
        try:
            rights_tab_btn = driver.find_element(By.CSS_SELECTOR, "#metaTab a[data-target='#meta-right']")
            rights_tab_btn.click()
        except Exception:
            pass
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "track-right-apply"))
        )
        # 1) 유통회원 검색 버튼 클릭 (열기)
        btn = None
        for locator in [
            (By.CSS_SELECTOR, "#track-right-apply button.search-group[data-group-category='S']"),
            (By.CSS_SELECTOR, "#track-right-apply button.search-group[name='search-grup2']"),
            (By.XPATH, "//div[@id='track-right-apply']//label[contains(.,'유통회원')]/following::button[contains(@class,'search-group')][1]")
        ]:
            try:
                btn = driver.find_element(*locator)
                if btn:
                    break
            except Exception:
                continue
        if btn is None:
            raise Exception("유통회원 검색 버튼을 찾을 수 없습니다.")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        try:
            btn.click()
        except ElementClickInterceptedException:
            driver.execute_script("arguments[0].click();", btn)

        # 2) 모달 표시 대기
        modal = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.ID, "rightModal"))
        )
        print("유통회원 검색 모달 열림")

        # 3) 검색어 입력 ('케이저')
        try:
            keyword_input = modal.find_element(By.CSS_SELECTOR, "input[type='text']")
        except Exception:
            keyword_input = modal.find_element(By.XPATH, ".//input[contains(translate(@name,'KEYWORD','keyword'),'keyword')]")
        driver.execute_script("arguments[0].focus();", keyword_input)
        try:
            keyword_input.clear()
        except Exception:
            pass
        keyword_input.send_keys("케이저")

        # 4) '검색' 버튼 클릭
        search_btn = None
        try:
            search_btn = modal.find_element(By.XPATH, ".//button[normalize-space(text())='검색']")
        except Exception:
            pass
        if search_btn is None:
            try:
                search_btn = modal.find_element(By.CSS_SELECTOR, "button.btn.btn-primary")
            except Exception:
                btns = modal.find_elements(By.TAG_NAME, "button")
                search_btn = btns[0] if btns else None
        if search_btn is None:
            raise Exception("검색 버튼을 찾을 수 없습니다.")
        try:
            search_btn.click()
        except ElementClickInterceptedException:
            driver.execute_script("arguments[0].click();", search_btn)

        # 5) 결과 테이블 로딩 대기 후 대상 선택 (stale 대응 재시도)
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#rightModal .modal-body, #rightModal table"))
        )

        def try_select_once() -> bool:
            try:
                # 모달을 매번 재조회하여 stale 방지
                modal_now = WebDriverWait(driver, 3).until(
                    EC.visibility_of_element_located((By.ID, "rightModal"))
                )
                rows_now = driver.find_elements(By.CSS_SELECTOR, "#rightModal .modal-body table tbody tr, #rightModal table tbody tr")
                if not rows_now:
                    return False
                # 대상 행 찾기
                target = None
                for tr in rows_now:
                    try:
                        tds = tr.find_elements(By.TAG_NAME, "td")
                        if any((cell.text.strip().lower().startswith("metalfocus")) for cell in tds):
                            target = tr
                            break
                        joined = " ".join([td.text.strip() for td in tds]).lower()
                        if "metalfocus" in joined:
                            target = tr
                            break
                    except StaleElementReferenceException:
                        return False
                if target is None:
                    return False
                # 클릭 대상 찾기
                clickable_local = None
                try:
                    clickable_local = target.find_element(By.XPATH, ".//button[contains(.,'선택') or contains(.,'선 정') or contains(.,'선 택')]")
                except Exception:
                    for sel in ["input[type='radio']", "input[type='checkbox']", "a", "button", "td"]:
                        try:
                            clickable_local = target.find_element(By.CSS_SELECTOR, sel)
                            break
                        except Exception:
                            continue
                if clickable_local is None:
                    clickable_local = target
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", clickable_local)
                try:
                    clickable_local.click()
                except ElementClickInterceptedException:
                    driver.execute_script("arguments[0].click();", clickable_local)
                print("유통회원: metalfocus* 항목 선택 완료")
                return True
            except StaleElementReferenceException:
                return False

        selected = False
        for _ in range(3):
            if try_select_once():
                selected = True
                break
            time.sleep(0.2)
        if not selected:
            raise Exception("유통회원 선택 중 요소 갱신으로 선택 실패")

        # 6) 선택 시 자동 닫힘 대기
        try:
            WebDriverWait(driver, 5).until(
                EC.invisibility_of_element_located((By.ID, "rightModal"))
            )
        except TimeoutException:
            # 안전 닫기
            try:
                close_btn = modal.find_element(By.CSS_SELECTOR, ".modal-footer button, .close")
                close_btn.click()
                WebDriverWait(driver, 3).until(
                    EC.invisibility_of_element_located((By.ID, "rightModal"))
                )
            except Exception:
                pass
    except Exception as e:
        print(f"유통회원 자동 선택 실패: {e}")

def goto_album_register(driver) -> bool:
    try:
        driver.get(ALBUM_REGISTER_URL)

        WebDriverWait(driver, 10).until(EC.url_contains("/mypage/meta"))
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "form, input, select, textarea"))
            )
        except Exception:
            pass

        print("앨범등록 페이지 진입 완료!")

        bulk_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "register-excel-btn"))
        )
        bulk_btn.click()
        excel_card = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "excel-card"))
        )
        WebDriverWait(driver, 10).until(lambda d: 'd-none' not in excel_card.get_attribute('class'))
        print("대량등록(엑셀) 패널 열기 완료!")

        file_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "mims-excel-upload"))
        )

        excel_filename = "박성태 - 기도왕｜MIMS.xlsx"
        excel_path = os.path.abspath(os.path.join(os.getcwd(), excel_filename))
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")

        driver.execute_script(
            "arguments[0].classList.remove('d-none'); arguments[0].style.display='block';",
            file_input,
        )
        file_input.send_keys(excel_path)
        print(f"파일 선택 완료: {excel_path}")

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#excel-card tbody tr"))
        )

        upload_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='excel-card']//button[.//span[contains(@class,'fa-upload')]]"))
        )
        upload_btn.click()
        print("업로드 버튼 클릭 완료. 업로드 진행 대기...")

        try:
            WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            print(f"업로드 경고창 감지: {alert.text}")
            alert.accept()
            print("경고창 확인(accept) 완료")
        except TimeoutException:
            pass

        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#excel-card .badge-success"))
            )
            print("업로드 완료 감지!")
        except TimeoutException:
            print("업로드 성공 배지를 확인하지 못했지만 다음 단계로 진행합니다.")

        try:
            search_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "search-btn"))
            )
            search_btn.click()
        except Exception:
            pass

        table_first_album_xpath = "//div[@id='table']//table//tbody/tr[1]/td[3]/a[contains(@class,'go-register')]"
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, table_first_album_xpath))
        )
        last_err = None
        for attempt in range(5):
            try:
                first_album_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, table_first_album_xpath))
                )
                first_album_text = first_album_link.text
                first_album_link.click()
                print(f"상세등록 페이지로 이동 중... (앨범명: {first_album_text})")
                break
            except StaleElementReferenceException as e:
                last_err = e
                time.sleep(0.5)
                continue
        else:
            raise last_err if last_err else Exception("첫 행 앨범 링크 클릭 실패")

        WebDriverWait(driver, 20).until(EC.url_contains("/mypage/meta/register/"))
        print("상세등록 페이지 진입 완료!")

        # 곡정보 탭이 보이지 않으면 이동 시도
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.ID, "track-list"))
            )
        except TimeoutException:
            _check_required_and_go_next(driver)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "track-list"))
            )

        # 업로드에 사용한 엑셀에서 재생시간을 읽어 각 트랙 HH/MM/SS 채우기
        try:
            _fill_durations_from_excel(driver, excel_path)
        except Exception as e:
            print(f"재생시간 채우기 실패: {e}")

        # 마지막 곡에서 '저장 후 다음으로'를 눌러 다음곡 없음 알림을 유도하고 확인 처리
        try:
            _save_next_on_last_track(driver)
        except Exception as e:
            print(f"마지막 곡 저장-다음 처리 실패: {e}")

        # 앨범중복확인 화면 분기 처리
        try:
            proceed = _handle_meta_confirm(driver)
            if not proceed:
                print("검색앨범 수록곡이 있어 자동 진행을 중단합니다.")
                return True
        except Exception as e:
            print(f"앨범중복확인 화면 분기 처리 실패: {e}")
            return True

        # 권리정보 탭에 진입한 경우, 제작회원/유통회원 검색 모달 띄우기
        try:
            if WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.ID, "right-reg-btn"))
            ):
                # 제작회원 자동 선택
                _select_producer_member(driver)
                # 유통회원: 오픈 → '케이저' 검색 → metalfocus 선택 (선택시 자동 닫힘)
                _select_distributor_member(driver)
                # 등록 버튼 클릭 및 완료 대기
                try:
                    reg_btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.ID, "right-reg-btn"))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", reg_btn)
                    reg_btn.click()
                    print("등록 버튼 클릭")
                    _drain_alerts_quick(driver)
                    # 등록 처리 후 My앨범으로 이동하거나 토스트/알럿이 있을 수 있어 잠시 대기
                    time.sleep(0.5)
                except Exception as e:
                    print(f"등록 버튼 클릭 실패: {e}")
        except Exception:
            pass

        # My앨범으로 이동하여 통합된 로직 실행 (가장 최신 앨범 대상으로 수행)
        try:
            albums = find_approved_albums(driver)
            if albums:
                latest_code = albums[0]["code"]
                album_url = f"https://www.mims.or.kr/mypage/view/album/{latest_code}"
                driver.get(album_url)
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, "//th[contains(text(), 'ISRC/Music.UCI')]/ancestor::table/tbody/tr"))
                )
                result = issue_codes(driver)
                if result:
                    print("main.py 로직 실행 완료 - 코드 결과:")
                    for item in result:
                        print(f"  곡명: {item['title']}")
                        print(f"  ISRC: {item['isrc']}")
                        print(f"  UCI:  {item['uci']}")
                        print("-" * 22)
                else:
                    print("main.py 로직 실행 결과: 코드 없음 또는 실패")
            else:
                print("My앨범에서 앨범을 찾지 못했습니다.")
        except Exception as e:
            print(f"My앨범 이동/코드 처리 실패: {e}")

        return True
    except Exception as e:
        print(f"앨범등록 페이지/업로드/상세 진입 흐름 실패: {e}")
        return False


def main():
    load_dotenv()
    mims_id = os.getenv("MIMS_ID")
    mims_password = os.getenv("MIMS_PASSWORD")

    if not mims_id or not mims_password:
        print("환경변수 MIMS_ID/MIMS_PASSWORD가 설정되지 않았습니다. .env를 확인하세요.")
        return

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    try:
        if not login(driver, mims_id, mims_password):
            return
        goto_album_register(driver)
        input("작업 완료. Enter를 누르면 브라우저를 닫습니다...")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
